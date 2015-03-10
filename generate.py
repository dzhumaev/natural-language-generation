# -*- coding: utf-8 -*-

from collections import defaultdict
from pprint import pprint

import openpyxl


import pymorphy2
morph = pymorphy2.MorphAnalyzer()


en_ru = {}


def to(word, *categories):
    return morph.parse(word)[0].inflect(set(categories)).word


def with_number_ru(number, word, *categories):
    return morph.parse(word)[0].inflect(set(categories)).make_agree_with_number(number).word


def with_number_en(number, word):
    return word + ('s' if number > 1 else '')


def make_player(ws, row_index):
    i = str(row_index)
    last_name_ru, first_name_ru = ws['E' + i].value.split()
    last_name_en, first_name_en = ws['C' + i].value.split()
    return {
        'first_name': {'en': first_name_en, 'ru': first_name_ru},
        'last_name': {'en': last_name_en, 'ru': last_name_ru},
        'role': {'en': ws['D' + i].value, 'ru': ws['F' + i].value},
    }

def load_logs(file_name):
    logs = []
    wb = openpyxl.load_workbook(file_name)
    for ws in wb:
        log = {}
        logs.append(log)
        log['date'] = ws['A2'].value
        log['arena'] = {'en': ws['B2'].value, 'ru': ws['B3'].value}
        log['city'] = {'en': ws['C2'].value, 'ru': ws['C3'].value}
        log['attendance'] = ws['D2'].value
        log['home-team'] = {'en': ws['E2'].value, 'ru': ws['E3'].value}
        log['guest-team'] = {'en': ws['F2'].value, 'ru': ws['F3'].value}
        en_ru[log['home-team']['en']] = log['home-team']['ru']
        en_ru[log['guest-team']['en']] = log['guest-team']['ru']
        log['players'] = defaultdict(dict)
        home_team_last_row = 4
        for i in range(4, 31):
            if ws['A'+str(i)].value == 'guest-team':
                home_team_last_row = i - 2
                break
        guest_team_last_row = home_team_last_row + 2
        for i in range(guest_team_last_row, 61):
            if ws['A'+str(i)].value == 'Play':
                guest_team_last_row = i - 2
                break
        for i in range(5, home_team_last_row+1):
            log['players'][log['home-team']['en']][ws['B'+str(i)].value] = make_player(ws, i)
        for i in range(home_team_last_row+3, guest_team_last_row+1):
            log['players'][log['guest-team']['en']][ws['B'+str(i)].value] = make_player(ws, i)

        row = ws.get_highest_row()
        log['score'] = [ws['B'+str(row-1)].value, ws['B'+str(row)].value]
        
        log['goals'] = []
        log['assists'] = []
        log['time'] = 'regulation'
        for i in range(guest_team_last_row+3, row-2):
            if ws['J'+str(i)].value in {'scored', 'powerplay', 'shorthanded'}:
                team = ws['D'+str(i)].value
                author = log['players'][team][ws['E'+str(i)].value]
                log['winning_goal'] = author
                if type(ws['B'+str(i)].value) == int:
                    log['goals'].append({
                        'minute': ws['B'+str(i)].value,
                        'team': team,
                        'author': author,
                        'type': ws['J'+str(i)].value,
                    })
                    if type(ws['H'+str(i)].value) == int:
                        log['assists'].append({
                            'minute': ws['B'+str(i)].value,
                            'team': team,
                            'author': log['players'][team][ws['H'+str(i)].value],
                        })
                    if type(ws['I'+str(i)].value) == int:
                        log['assists'].append({
                            'minute': ws['B'+str(i)].value,
                            'team': team,
                            'author': log['players'][team][ws['I'+str(i)].value],
                        })

            if ws['A'+str(i)].value == 'End of overtime':
                log['time'] = 'overtime'
            if ws['A'+str(i)].value == 'Shootout':
                log['time'] = 'shootout'
                
        process_log(log)
    return logs


def process_log(log):
    if log['score'][0] > log['score'][1]:
        log['winner'], log['loser'] = log['home-team'], log['guest-team']
    else:
        log['loser'], log['winner'] = log['home-team'], log['guest-team']
    log['score'] = '-'.join(map(str, sorted(log['score'], reverse=True)))


class Event:
    def __init__(self, log):
        self.log = log

    def is_applicable(self):
        return True

    def wrap(self, message):
        return message

    def gen_wrapped_russian(self):
        return self.wrap(self.gen_russian())

    def gen_wrapped_english(self):
        return self.wrap(self.gen_english())

    def gen_russian(self):
        return ''

    def gen_english(self):
        return ''


class Pause(Event):
    def wrap(self, message):
        return '<break time="1000ms"/>'


class HighPitchEvent(Event):
    # def wrap(self, message):
    #     return '<prosody rate="slow" pitch="high">' + message + '</prosody>'
    pass


class WinnerEvent(HighPitchEvent):
    def gen_russian(self):
        time = {'regulation': ' в основное время', 'overtime': ' в овертайме',
                'shootout': ' по буллитам'}
        return [self.log['winner']['ru'] + ' обыграл ' + self.log['loser']['ru']
                + time[self.log['time']] + ' со счётом ' + self.log['score']]

    def gen_english(self):
        time = {'regulation': ' in regulation', 'overtime': ' in overtime',
                'shootout': ' in shootout'}
        return [self.log['winner']['en'] + ' picked up a ' + self.log['score']
                + ' win against ' + self.log['loser']['en'] + time[self.log['time']]]


def join_with_and(words, and_word):
    return ', '.join(words[:-1]) + ' ' + and_word + ' ' + words[-1]


def join_with_and_ru(words):
    return join_with_and(words, 'и')


def join_with_and_en(words):
    return join_with_and(words, 'and')


class GoalsSummaryEvent(HighPitchEvent):
    def __init__(self, log):
        super(GoalsSummaryEvent, self).__init__(log)
        self.goals_by_team_ru = defaultdict(set)
        self.goals_by_team_en = defaultdict(set)
        for goal in log['goals']:
            self.goals_by_team_ru[goal['team']].add(goal['author']['last_name']['ru'])
            self.goals_by_team_en[goal['team']].add(goal['author']['last_name']['en'])
        for k, v in self.goals_by_team_ru.items():
            self.goals_by_team_ru[k] = sorted(v)
        for k, v in self.goals_by_team_en.items():
            self.goals_by_team_en[k] = sorted(v)

    def gen_russian(self):
        winning = {'regulation': '', 'overtime': ', ',
                  'shootout': ', ' + self.log['winning_goal']['last_name']['ru']
                  + ' забросил победный буллит'}
        return ['В составе ' + self.log['winner']['ru'] + ' отличились '
                + join_with_and_ru(self.goals_by_team_ru[self.log['winner']['en']]) 
                + winning[self.log['time']]
                + ', а за ' + self.log['loser']['ru'] + ' шайбы забили '
                + join_with_and_ru(self.goals_by_team_ru[self.log['loser']['en']])]

    def gen_english(self):
        winning = {'regulation': '', 'overtime': ', ',
                  'shootout': ', ' + self.log['winning_goal']['last_name']['en']
                  + ' scored the winning bullet'}
        return [join_with_and_en(self.goals_by_team_en[self.log['winner']['en']])
                + ' were on target for ' + self.log['winner']['en']
                + winning[self.log['time']] + ', while '
                + join_with_and_en(self.goals_by_team_en[self.log['loser']['en']])
                + ' replied for ' + self.log['loser']['en']]


ORDINAL_RU = ['первый', 'второй', 'третий']
ORDINAL_EN = ['opening', 'second', 'third']


def join_sentences(sentences):
    return ' '.join(sentence[:1].upper() + sentence[1:] + '.' for sentence in sentences)


def say_player_ru(player, team):
    return player['role']['ru'] + ' ' + to(en_ru[team], 'gent').title() + ' ' + player['last_name']['ru']


def say_player_en(player, team):
    return player['first_name']['en'] + ' ' + player['last_name']['en']


class GoalsByPeriodEvent(Event):
    def __init__(self, log):
        super(GoalsByPeriodEvent, self).__init__(log)
        periods = self.periods = [defaultdict(list) for i in range(4)]
        for goal in log['goals']:
            if goal['minute'] <= 20:
                period_index = 0
            elif goal['minute'] <= 40:
                period_index = 1
            elif goal['minute'] <= 60:
                period_index = 2
            else:
                period_index = 3
            periods[period_index][goal['team']].append(goal)

    def gen_russian(self):
        chunks = []
        for i, period in enumerate(self.periods[:3]):
            if not period:
                chunks.append(ORDINAL_RU[i] + ' период оказался нерезультативным')
            elif len(period) == 1:
                team, goals = list(period.items())[0]
                prefix = 'в ' + to(ORDINAL_RU[i], 'loct') + ' периоде '
                if len(goals) == 1:
                    goal = goals[0]
                    if goal['minute'] % 20 >= 15:
                        prefix = 'под конец ' + to(ORDINAL_RU[i], 'gent') + ' периода '
                    chunks.append(
                        prefix + say_player_ru(goal['author'], team) + ' забил одну шайбу'
                    )
                else:
                    num_goals = len(goals)
                    chunks.append(
                        prefix + 'удача была на стороне игроков ' + en_ru[team]
                        + ', которые забросили ' + str(num_goals) + ' '
                        + with_number_ru(num_goals, 'шайбу')
                    )
            else:
                (team1, goals1), (team2, goals2) = sorted(period.items(), key=lambda x: len(x[1]))
                num_goals1 = len(goals1)
                num_goals2 = len(goals2)
                chunks.append(
                    'в ' + to(ORDINAL_RU[i], 'loct') + ' периоде ' + en_ru[team1] + ' '
                    + str(num_goals1) + ' ' + with_number_ru(num_goals1, 'раз')
                    + ' поразил ворота соперников, ' + en_ru[team2] + ' ответил ' + str(num_goals2) + ' '
                    + with_number_ru(num_goals2, 'забитой', 'ablt') + ' '
                    + with_number_ru(num_goals2, 'шайбой')
                )

        return chunks

    def gen_english(self):
        chunks = []
        for i, period in enumerate(self.periods[:3]):
            suffix = ' in the ' + ORDINAL_EN[i] + ' period'
            if not period:
                chunks.append('there were no goals scored' + suffix)
            elif len(period) == 1:
                team, goals = list(period.items())[0]
                if len(goals) == 1:
                    goal = goals[0]
                    if goal['minute'] % 20 >= 15:
                        suffix = ' in the end of the ' + ORDINAL_EN[i] + ' period'
                    chunks.append(
                        say_player_en(goal['author'], team) + ' scored one goal' + suffix
                    )
                else:
                    num_goals = len(goals)
                    chunks.append(
                        team + ' scored ' + str(num_goals) + ' '
                        + with_number_en(num_goals, 'goal') + suffix
                    )
            else:
                (team1, goals1), (team2, goals2) = sorted(period.items(), key=lambda x: len(x[1]))
                num_goals1 = len(goals1)
                num_goals2 = len(goals2)
                chunks.append(
                    team1 + ' took the puck to the net ' + str(num_goals1) + ' '
                    + with_number_en(num_goals1, 'time') + suffix + ', while ' + team2 + ' responded with '
                    + str(num_goals2) + ' ' + with_number_en(num_goals2, 'goal')
                )

        return chunks

EVENT_CLASSES = [WinnerEvent,
                 GoalsSummaryEvent,
                 GoalsByPeriodEvent,
                 ]


def form_report(log):
    russian_report_chunks = []
    english_report_chunks = []

    for event_class in EVENT_CLASSES:
        event = event_class(log)
        if event.is_applicable():
            russian_report_chunks.extend(event.gen_wrapped_russian())
            english_report_chunks.extend(event.gen_wrapped_english())

    reports = [join_sentences(russian_report_chunks), join_sentences(english_report_chunks)]
    return [''.join(report) for report in reports]


def main():
    logs = load_logs('Hockey_Log.xlsx')
    pprint(logs)
    for log in logs:
        for report in form_report(log):
            print(report)


if __name__ == '__main__':
    main()
